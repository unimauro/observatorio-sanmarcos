import openpyxl
wb=openpyxl.load_workbook("conosce/CONOSCE_ADJUDICACIONES2023_0.xlsx", read_only=True)
ws=wb[wb.sheetnames[0]]
rows=ws.iter_rows(values_only=True)
hdr=next(rows)
for i,h in enumerate(hdr):
    print(i, h)
print("--- first data row ---")
r=next(rows)
for i,v in enumerate(r):
    print(i, repr(v))
wb.close()
