import openpyxl, json, os
HERE=os.path.dirname(os.path.abspath(__file__))
CODIGO='001906'; RUC_ENT='20148092282'
def tp(ruc):
    p=str(ruc)[:2]
    return 'natural' if p in ('10','15') else 'empresa'
duenos={
 '20100070970':("Grupo Intercorp (Super Food Holding); controlado por Carlos Rodriguez-Pastor Persivale. Opera Plaza Vea, Vivanda, Mass, Makro",
                "https://www.datosperu.org/empresa-supermercados-peruanos-sociedad-anonima-o-spsa-20100070970.php"),
 '20507634479':("Edenred SE (Francia, cotiza en Euronext Paris; ex-Accor Services). Presidente/CEO Bertrand Dumazy. Emite vales tipo Ticket Restaurant",
                "https://en.wikipedia.org/wiki/Edenred"),
 '20552504641':("Grupo Optical Networks; fundador Eduardo Barriga. Division residencial (Wi-Net) vendida a Linzor Capital en 2024; Win Empresas sigue en el grupo. GG Eduardo Zagazeta",
                "https://www.infobae.com/peru/2024/08/22/venden-win-linzor-capital-comprara-la-empresa-peruana-pero-solo-una-parte-cambiara-de-dueno/"),
 '20502797230':("Diveimport S.A. (marca Divemotor); parte del Grupo Kaufmann (Chile), distribuidor de Mercedes-Benz, Jeep, Freightliner",
                "https://www.ey.com/es_pe/insights/revista-execution/entrevistas/entrevista-divemotor-peru"),
 '20502257987':("Corporacion Vega (Tiendas Vega); familia Vega. Gerente general Antonio Michel Vega Paredes",
                "https://www.datosperu.org/empresa-corporacion-vega-sac-20502257987.php"),
 '20423044765':("Gerente general Carlos Alberto Azcarate Lucero (desde 2002)",
                "https://www.datosperu.org/empresa-corporacion-alimentaria-alexcar-sac-20423044765.php"),
 '20106897914':("Entel Peru S.A., filial del grupo chileno Entel (Empresa Nacional de Telecomunicaciones S.A.)",
                "https://es.wikipedia.org/wiki/Entel_Chile"),
}
rows=[]
for Y in (2023,2024,2025):
    wb=openpyxl.load_workbook(f"conosce/CONOSCE_ADJUDICACIONES{Y}_0.xlsx", read_only=True)
    ws=wb[wb.sheetnames[0]]
    it=ws.iter_rows(values_only=True); next(it)
    for r in it:
        if str(r[0])==CODIGO and str(r[1])==RUC_ENT:
            rows.append(r)
    wb.close()
print("item rows:", len(rows))

agg={}
for r in rows:
    ruc=str(r[19]); monto=r[15] or 0
    a=agg.setdefault(ruc,{'nombre':r[20],'ruc':ruc,'monto':0.0,'convs':set(),
        'tipos':{},'objeto':{},'tipo_persona':tp(ruc),'tipo_proveedor_seace':r[21]})
    a['monto']+=float(monto)
    a['convs'].add(r[5])
    a['tipos'][r[7]]=a['tipos'].get(r[7],0)+1
    a['objeto'][r[6]]=a['objeto'].get(r[6],0)+1
provs=[]
for ruc,a in agg.items():
    d,f=duenos.get(ruc,(None,None))
    provs.append({'nombre':a['nombre'],'ruc':ruc,'monto':round(a['monto'],2),
        'n':len(a['convs']),'tipos':a['tipos'],'objeto':a['objeto'],
        'tipo_persona':a['tipo_persona'],'tipo_proveedor_seace':a['tipo_proveedor_seace'],
        'dueno':d,'fuente_dueno':f})
provs.sort(key=lambda x:-x['monto'])
monto_total=round(sum(p['monto'] for p in provs),2)
emp=[p for p in provs if p['tipo_persona']=='empresa']
nat=[p for p in provs if p['tipo_persona']=='natural']
all_convs=set(r[5] for r in rows)
top_personas=sorted([{'nombre':p['nombre'],'ruc':p['ruc'],'monto':p['monto'],'n':p['n']} for p in nat],key=lambda x:-x['monto'])
out={
 '_meta':{
  'fuente':'OECE/OSCE - CONOSCE Datos Abiertos, reporte de Adjudicaciones (buena pro por item)',
  'fuente_url':'https://conosce.osce.gob.pe/buscador/assets/67ae6c4a/reportes/adjudicaciones/',
  'entidad':'Universidad Nacional Mayor de San Marcos (UNMSM) - pliego 510',
  'ruc':RUC_ENT,
  'codigoentidad_conosce':int(CODIGO),
  'periodo':'2023-2025',
  'extraido':'2026-07',
  'unidad_monto':'Soles (PEN), monto adjudicado por item',
  'nota':"Agregado desde reportes anuales CONOSCE de Adjudicaciones (nivel item de buena pro), archivos CONOSCE_ADJUDICACIONES{2023,2024,2025}_0.xlsx, filtrando codigoentidad 001906 / RUC 20148092282 (UNIVERSIDAD NACIONAL MAYOR DE SAN MARCOS, entidad de Gobierno Nacional; NO confundir con las municipalidades 'San Marcos'). 'monto' = suma de monto_adjudicado_item_soles. 'n' = numero de procesos (codigoconvocatoria) distintos en que ese proveedor obtuvo buena pro. 'tipos' = procesos por tipo de proceso de seleccion. 'objeto' = items por objeto contractual (Bien/Servicio/Obra). tipo_persona por prefijo de RUC (20=empresa, 10/15=natural); consorcios y no domiciliados clasificados como empresa. NO incluye ordenes de compra <8 UIT (dataset aparte)."
 },
 'totales':{
  'monto_total':monto_total,
  'n_proveedores':len(provs),
  'n_procesos':len(all_convs),
  'n_empresas':len(emp),
  'n_personas_naturales':len(nat),
  'monto_empresas':round(sum(p['monto'] for p in emp),2),
  'monto_personas_naturales':round(sum(p['monto'] for p in nat),2),
 },
 'top_personas':top_personas,
 'proveedores':provs,
}
json.dump(out, open(os.path.join(HERE,'..','data','proveedores-sanmarcos.json'),'w'),
          ensure_ascii=False, indent=1)
print("proveedores:", len(provs), "monto_total:", monto_total, "procesos:", len(all_convs))
print("empresas:", len(emp), "naturales:", len(nat))
print("--- TOP 15 ---")
for p in provs[:15]:
    print(f"  {p['monto']:>14,.2f}  {p['ruc']}  {p['nombre']}")
