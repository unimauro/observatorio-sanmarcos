import openpyxl, unicodedata
from collections import Counter
def norm(s):
    if s is None: return ''
    s=''.join(c for c in unicodedata.normalize('NFD',str(s)) if unicodedata.category(c)!='Mn')
    return s.upper()
for Y in (2023,2024,2025):
    f=f"conosce/CONOSCE_ADJUDICACIONES{Y}_0.xlsx"
    wb=openpyxl.load_workbook(f, read_only=True)
    ws=wb[wb.sheetnames[0]]
    rows=ws.iter_rows(values_only=True); next(rows)
    hits=Counter()
    for r in rows:
        ent=norm(r[2])
        if 'SAN MARCOS' in ent:
            hits[(r[0], r[1], r[2], norm(r[3]))]+=1
    print(f"--- {Y} ---")
    for k,v in sorted(hits.items(), key=lambda x:-x[1]):
        print(v, k)
    wb.close()
